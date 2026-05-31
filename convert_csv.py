import streamlit as st
import pandas as pd
import io
import openpyxl
from datetime import datetime, date
from decimal import Decimal
import re

st.set_page_config(page_title="Excel → CSV", page_icon="📊", layout="centered")

st.title("📊 Conversor Excel → CSV")
st.markdown("Carga un archivo Excel y descárgalo como CSV separado por coma, **sin modificar el formato de los datos**.")

# --- Funciones de utilidad ---

def contar_decimales_formato(fmt):
    """
    Extrae cuántos decimales indica el número_format de Excel.
    Ej: '0.00' → 2, '#,##0.000' → 3, '0' → 0, 'General' → None
    """
    if not fmt or fmt.strip().upper() in ("GENERAL", "@", ""):
        return None

    # Ignorar secciones de formato negativo/cero/texto (separadas por ;)
    seccion = fmt.split(";")[0]

    # Quitar secciones de color/condición como [Red], [$€-...]
    seccion = re.sub(r'\[.*?\]', '', seccion)

    # Buscar la parte decimal: dígitos después del punto en el patrón numérico
    match = re.search(r'\.([0#?]+)', seccion)
    if match:
        return len(match.group(1))
    return 0


def numero_a_str(valor, fmt):
    """
    Convierte un número a string respetando el formato visual de Excel.
    - Si el formato indica N decimales → aplica exactamente N decimales.
    - Si el formato es General o no reconocido → usa la representación natural.
    """
    try:
        decimales = contar_decimales_formato(fmt)

        if decimales is None:
            # Formato General: representación natural sin ceros innecesarios
            d = Decimal(str(valor)).normalize()
            _, _, exp = d.as_tuple()
            if exp >= 0:
                return str(int(valor))
            return format(d, 'f')

        # Formato explícito: respetar cantidad de decimales
        return f"{valor:.{decimales}f}"

    except Exception:
        return str(valor)


def celda_a_str(celda):
    """Convierte una celda openpyxl a string respetando fechas y formato numérico."""
    valor = celda.value
    fmt = celda.number_format if celda.number_format else "General"

    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d-%m-%Y")
    if isinstance(valor, date):
        return valor.strftime("%d-%m-%Y")
    if isinstance(valor, (int, float)):
        return numero_a_str(valor, fmt)
    return str(valor)


def leer_hoja(ws):
    """Lee la hoja celda a celda accediendo al objeto celda completo (no values_only)."""
    filas = []
    for fila in ws.iter_rows():
        filas.append([celda_a_str(celda) for celda in fila])
    return filas


def filas_a_csv_bytes(filas):
    output = io.StringIO()
    for fila in filas:
        celdas_escapadas = []
        for celda in fila:
            if ',' in celda or '"' in celda or '\n' in celda:
                celda = '"' + celda.replace('"', '""') + '"'
            celdas_escapadas.append(celda)
        output.write(",".join(celdas_escapadas) + "\n")
    return output.getvalue().encode("utf-8-sig")


# --- UI ---

archivo = st.file_uploader("Selecciona un archivo Excel", type=["xlsx", "xls", "xlsm"])

if archivo:
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        hojas = wb.sheetnames

        hoja_sel = st.selectbox("Selecciona la hoja a convertir:", hojas)
        ws = wb[hoja_sel]

        filas = leer_hoja(ws)
        n_filas = len(filas)
        n_cols = max((len(f) for f in filas), default=0)

        st.markdown(f"**Hoja:** `{hoja_sel}` · {n_filas} filas × {n_cols} columnas")

        if filas:
            preview_df = pd.DataFrame(filas[1:11], columns=filas[0] if filas else [])
            st.markdown("**Vista previa (primeras 10 filas):**")
            st.dataframe(preview_df, use_container_width=True)

        csv_bytes = filas_a_csv_bytes(filas)
        nombre_csv = archivo.name.rsplit(".", 1)[0] + f"_{hoja_sel}.csv"

        st.download_button(
            label="⬇️ Descargar CSV",
            data=csv_bytes,
            file_name=nombre_csv,
            mime="text/csv",
        )

        st.success(f"Archivo listo: **{nombre_csv}** ({len(csv_bytes):,} bytes)")

    except Exception as e:
        st.error(f"Error al procesar el archivo: {e}")

st.markdown("---")
st.caption("Fechas en dd-mm-yyyy · Decimales según formato visual de Excel · Textos y vacíos sin transformación.")
